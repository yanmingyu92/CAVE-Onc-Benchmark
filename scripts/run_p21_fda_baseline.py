"""Empirical Pinnacle 21 Community *FDA production engine* baseline on Track B.

The paper's original R1-3 response argued that Pinnacle 21's own production engines
(FDA/PMDA) would miss the non-CORE cross-domain RECIST contradictions, but only ran the
open CORE engine. This script runs the **real branded P21 FDA engine** on the *same*
`bench/core_lean/{clean,A01..A20}` corpora the CORE baseline used, so the two baselines
are directly comparable.

Pipeline
--------
``run``     : invoke the P21 Community CLI (FDA 2405.2, SDTMIG 3.4) on the clean corpus
              and each of the 20 injected corpora -> one .xlsx report each.
``analyze`` : diff each injected report's issue set against the clean report, adjudicate
              per-archetype detection, and emit eval/p21_fda_benchmark.json (mirroring
              eval/core_p21_benchmark.json) plus eval/p21_fda_new_issues.json (raw delta
              for transparent adjudication).
``all``     : run then analyze.

Detection criterion (mirrors scripts/run_core_p21_baseline.py):
an archetype is "detected by P21" iff a NEW P21 issue whose rule semantics ARE the
injected contradiction ("direct") appears in the injected report but not the clean one.
The FDA engine is domain-scoped/structural, so cross-domain RECIST contradictions are
expected to produce no such new issue.
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import subprocess
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parents[1]
CORPUS_DIR = ROOT / "bench" / "core_lean"
REPORT_DIR = ROOT / "eval" / "p21_fda_reports"

# Pinnacle 21 Community CLI (branded FDA engine). Java 8 is bundled with the install.
P21_JAVA = Path(
    r"C:\Users\yanmi\AppData\Local\Programs\Pinnacle 21 Community"
    r"\resources\app.asar.unpacked\components\java64\bin\java.exe"
)
P21_JAR = Path(
    r"C:\Users\yanmi\AppData\Local\Programs\Pinnacle 21 Community"
    r"\resources\app.asar.unpacked\components\lib\p21-client-1.0.8.jar"
)
# FDA 2405.2 is the current engine; the legacy 1903.1 config has expired (IqException).
P21_ENGINE = "FDA 2405.2"
P21_STANDARD_VERSION = "3.4"

ARCHETYPES = [f"A{i:02d}" for i in range(1, 21)]


# -- running the P21 CLI ------------------------------------------------------

def _run_one(corpus: str, force: bool = False) -> Path:
    """Validate one corpus dir with the P21 FDA engine -> .xlsx report path."""
    src = CORPUS_DIR / corpus
    if not src.is_dir():
        raise FileNotFoundError(f"corpus dir missing: {src} (run run_core_p21_baseline build)")
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report = REPORT_DIR / f"{corpus}.xlsx"
    if report.is_file() and not force:
        logger.info("skip %s (report exists)", corpus)
        return report
    cmd = [
        str(P21_JAVA), "-jar", str(P21_JAR),
        f"--engine.version={P21_ENGINE}",
        "--standard=sdtm",
        f"--standard.version={P21_STANDARD_VERSION}",
        f"--source.sdtm={src}",
        f"--report={report}",
    ]
    logger.info("P21 validate: %s", corpus)
    # P21 returns a non-zero exit when it finds Reject-severity issues; the report is
    # still written. Treat "report produced" as success rather than trusting returncode.
    subprocess.run(cmd, capture_output=True, text=True)
    if not report.is_file():
        raise RuntimeError(f"P21 produced no report for {corpus}")
    return report


def run(force: bool = False) -> None:
    _run_one("clean", force=force)
    for aid in ARCHETYPES:
        if (CORPUS_DIR / aid).is_dir():
            _run_one(aid, force=force)
        else:
            logger.warning("no injected corpus for %s (skipped)", aid)


# -- parsing the P21 .xlsx report ---------------------------------------------

# GLOBAL "missing dataset / missing define.xml" notices are corpus-composition noise,
# not data contradictions; they are identical across clean/injected and cancel in the
# delta, but we drop them explicitly so the raw new-issue dump stays readable.
_NOISE_IDS = {"DD0101"}


def parse_report(xlsx: Path) -> list[dict[str, Any]]:
    """Return the 'Details' sheet issue rows as normalized dicts."""
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if "Details" not in wb.sheetnames:
        return []
    ws = wb["Details"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    def g(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    out: list[dict[str, Any]] = []
    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        rid = g(row, "Pinnacle 21 ID")
        if rid in _NOISE_IDS:
            continue
        out.append({
            "domain": (g(row, "Domain") or ""),
            "record": g(row, "Record"),
            "count": g(row, "Count"),
            "variables": str(g(row, "Variables") or ""),
            "values": str(g(row, "Values") or ""),
            "rule_id": (rid or ""),
            "message": str(g(row, "Message") or "")[:160],
            "category": str(g(row, "Category") or ""),
            "severity": str(g(row, "Severity") or ""),
        })
    return out


def _issue_key(row: dict) -> tuple:
    """Signature used to diff clean vs injected (excludes volatile Count/Record)."""
    return (row["domain"], row["rule_id"], row["variables"], row["values"],
            row["message"])


# -- adjudication -------------------------------------------------------------

def _load_core_xref() -> dict[str, str | None]:
    """docs/frozen_archetype_list.csv -> {archetype: CORE-seed or None}."""
    path = ROOT / "docs" / "frozen_archetype_list.csv"
    out: dict[str, str | None] = {}
    with path.open(encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            x = (r.get("core_xref") or "").strip()
            out[r["archetype_id"]] = x if x.startswith("CORE-") else None
    return out


# Per-archetype adjudication of the P21 FDA engine's NEW issues (built empirically from
# eval/p21_fda_new_issues.json after the first run; see analyze()). Same semantics as the
# CORE CLASSIFICATION:
#   (matched_p21_rule, confidence, rationale)
#   confidence "direct"   = the P21 rule's semantics ARE the injected contradiction
#              "indirect" = flagged only via a generic derived-variable rule any edit trips
#              "none"     = P21 silent, or only collateral flags unrelated to the contradiction
# NOTE: filled in P4 after inspecting the raw deltas. Until then every archetype scores
# "none" and the headline reflects only what P21 demonstrably fires on.
P21_CLASSIFICATION: dict[str, tuple[str | None, str, str]] = {
    "A01": (None, "none", "RSORRES=PD vs >=30% SLD decrease; P21 silent (0 new issues)"),
    "A02": (None, "none", "P21 fires only SD0052 (collateral VISITNUM-within-VISIT label inconsistency at the new-lesion visit), not the new-lesion-vs-SD RECIST contradiction"),
    "A03": (None, "none", "P21 fires only collateral EX date errors (SD0013 EXSTDTC>EXENDTC, SD1090 EXSTDY miscalc), not the exposure-after-attributed-AE temporal contradiction"),
    "A04": (None, "none", "P21 fires only SD0052 (collateral intra-domain VISITNUM/VISIT label inconsistency), not the cross-domain visit-window propagation"),
    "A05": (None, "none", "SUPPDM SEX vs DM SEX conflict; P21 silent (0 new issues)"),
    "A06": (None, "none", "non-target PD vs RSORRES=SD; P21 silent (0 new issues)"),
    "A07": (None, "none", "PR without 28d confirmation; P21 silent (0 new issues)"),
    "A08": ("SD0066", "direct", "SD0066 'Invalid ARMCD' (ARMCD not in TA codelist) IS the injected contradiction"),
    "A09": ("SD1374", "direct", "SD1374 'No Disposition record found for subject' IS the contradiction"),
    "A10": (None, "none", "RFXSTDTC != earliest EXSTDTC; P21 FDA has no reference-date aggregation rule -> silent (CORE detects this)"),
    "A11": (None, "none", "RFXENDTC != latest EX date; P21 silent, no such rule (CORE detects this)"),
    "A12": ("SD2005", "direct", "SD2005 'Missing DTHDTC when DTHFL populated' IS the death-flag contradiction"),
    "A13": (None, "none", "RACE=MULTIPLE w/o SUPPDM; P21 silent, no such rule (CORE detects this)"),
    "A14": ("SD0070", "direct", "SD0070 'No Exposure record found for subject' IS the contradiction (CORE missed this)"),
    "A15": ("SD1086", "indirect", "RFSTDTC shift surfaced only via SD1086/SD1090/SD1094 study-day (DMDY/AESTDY/DSSTDY) miscalc -- generic derived-variable rules any date edit trips; specificity to the contradiction not shown"),
    "A16": ("SD1349", "direct", "SD1349 'Inconsistent STUDYID' for the duplicated USUBJID IS the duplicate-across-studies contradiction"),
    "A17": ("SD0071", "direct", "SD0071 'Invalid ARM/ARMCD' + SD1034 -- ARMCD-ARM not 1:1 IS the contradiction"),
    "A18": (None, "none", "non-target PD vs RS=SD; P21 silent (0 new issues)"),
    "A19": (None, "none", "RSORRES=CR vs target PR (Table 7); P21 silent (0 new issues)"),
    "A20": (None, "none", "P21 fires only collateral SD0005 (duplicate RSSEQ) + SD0051/SD0052 (VISIT label inconsistency), not the iRECIST iCPD temporal contradiction"),
}


def analyze() -> dict:
    clean = parse_report(REPORT_DIR / "clean.xlsx")
    clean_keys = {_issue_key(r) for r in clean}
    xref = _load_core_xref()

    new_issues_dump: dict[str, list[dict]] = {}
    results = []
    direct = indirect = flags_any = 0
    sem_total = sem_direct = 0
    struct_total = struct_direct = struct_indirect = 0
    for aid in ARCHETYPES:
        rep = REPORT_DIR / f"{aid}.xlsx"
        if not rep.is_file():
            results.append({"archetype": aid, "error": "no P21 report"})
            continue
        inj = parse_report(rep)
        new = [r for r in inj if _issue_key(r) not in clean_keys]
        new_issues_dump[aid] = new
        new_rule_ids = sorted({r["rule_id"] for r in new})

        matched_rule, confidence, rationale = P21_CLASSIFICATION.get(
            aid, (None, "none", "unadjudicated"))
        design_core_seed = xref.get(aid)
        category = ("core_derived_structural" if design_core_seed
                    else "noncore_crossdomain")
        rule_fired = bool(matched_rule) and matched_rule in new_rule_ids
        is_direct = rule_fired and confidence == "direct"
        is_indirect = rule_fired and confidence == "indirect"

        flags_any += int(len(new) > 0)
        direct += int(is_direct)
        indirect += int(is_indirect)
        if category == "noncore_crossdomain":
            sem_total += 1
            sem_direct += int(is_direct)
        else:
            struct_total += 1
            struct_direct += int(is_direct)
            struct_indirect += int(is_indirect)

        results.append({
            "archetype": aid,
            "category": category,
            "design_core_seed": design_core_seed,
            "p21_new_issue_count": len(new),
            "p21_flags_any": len(new) > 0,
            "detection_confidence": confidence if rule_fired else "none",
            "p21_detects_contradiction": is_direct,
            "matched_p21_rule": matched_rule if rule_fired else None,
            "new_p21_rule_ids": new_rule_ids,
            "rationale": rationale,
        })

    n = sum(1 for r in results if "error" not in r)
    summary = {
        "engine": f"Pinnacle 21 Community — branded FDA production engine ({P21_ENGINE})",
        "standard": f"sdtmig {P21_STANDARD_VERSION}",
        "corpus": "Track B lean benchmark (same bench/core_lean corpora as the CORE baseline)",
        "clean_issues": len(clean),
        "archetypes_total": n,
        "detection_criterion": (
            "P21 detects an archetype iff a NEW P21 issue whose rule semantics ARE the "
            "injected contradiction ('direct') appears vs clean. 'indirect' = generic "
            "derived-variable rule any edit trips; 'collateral'/silence = not detected."),
        "HEADLINE": {
            "p21_detects_DIRECT": f"{direct}/{n}",
            "p21_detects_direct_plus_indirect": f"{direct + indirect}/{n}",
            "p21_flags_subject_any_reason": f"{flags_any}/{n}",
            "noncore_crossdomain_detected": f"{sem_direct}/{sem_total}",
            "core_derived_structural_detected_direct": f"{struct_direct}/{struct_total}",
            "core_derived_structural_detected_incl_indirect":
                f"{struct_direct + struct_indirect}/{struct_total}",
        },
        "note": (
            "Real branded Pinnacle 21 FDA engine on the Track B corpus. Compare against "
            "eval/core_p21_benchmark.json (open CORE engine). Both are domain-scoped "
            "structural validators; the non-CORE cross-domain RECIST class is the "
            "expressiveness gap CAVE targets."),
        "per_archetype": results,
    }
    (ROOT / "eval" / "p21_fda_new_issues.json").write_text(
        json.dumps(new_issues_dump, indent=2, default=str), encoding="utf-8")
    out = ROOT / "eval" / "p21_fda_benchmark.json"
    out.write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")
    logger.info("P21 FDA direct %d/%d | non-CORE %d/%d | structural %d/%d -> %s",
                direct, n, sem_direct, sem_total, struct_direct, struct_total, out)
    return summary


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("command", choices=["run", "analyze", "all"])
    ap.add_argument("--force", action="store_true", help="re-run P21 even if reports exist")
    args = ap.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    if args.command in ("run", "all"):
        run(force=args.force)
    if args.command in ("analyze", "all"):
        summary = analyze()
        print(json.dumps({k: v for k, v in summary.items() if k != "per_archetype"},
                         indent=2, default=str))


if __name__ == "__main__":
    main()
